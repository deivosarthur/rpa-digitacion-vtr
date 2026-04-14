# Proyecto : RPA "TG Bots"
# Propietario : Technogroup SPA
# Creador: Jose Huaiquiman Arce

from selenium import webdriver  #
import time  #Ejercicio 1 , libreria nativa de python, para controlar el tiempo de espera del time.sleep()
from selenium.webdriver.common.by import By #
from selenium.webdriver.common.keys import Keys #
from selenium.webdriver.support.ui import WebDriverWait #
from selenium.webdriver.support import expected_conditions as EC #
import unittest #
import cv2    # se debe instalar con "pip install opencv-python" desde cmd como administrador
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains #permite importar la clase ActionChains, esta clase permite automatizar interacciones de usuario más complejas en una página web
from selenium.webdriver.chrome.options import Options #permite personalizar el comportamiento del navegador Chrome antes de que se inicie
from openpyxl import load_workbook  # La poder leer archivos de excel
import os  # Permite ver que carpeta esta direccionada la busqueda de un archivo
import csv # para manejo de archivos csv
import re # ¡Añade esta línea!
import sys
import json
from pathlib import Path


#from selenium.common.exceptions import TimeoutException
#from playsound import playsound # primero se ejecuta (pip install pyglet) y luego (pip install playsound)
#print(os.getcwd()) # Muestra la ruta de archivos adicionales

monitor="4"
archivolog=f"./Digitacion/log_ejecucion_monitor{monitor}.csv"
statusorden=f"./Digitacion/status_orden_de_trabajo.csv"
os.makedirs(os.path.dirname(archivolog), exist_ok=True)

class usando_unittest(unittest.TestCase):
    
    def setUp(self):
        self.driver = webdriver.Chrome()
        self.driver.set_window_rect(x=450, y=0, width=800, height=800)
        self.driver.implicitly_wait(10) # Espera implícita para evitar errores de elementos no encontrados.

    def test_a_run_1_de_2(self):
        driver = self.driver
        driver.get("http://eps.vtr.cl/eps/declaracion_consumo.do#")
        driver.execute_script("document.body.style.zoom='65%'")
        with open( archivolog, 'a', newline='') as csvfile: # 'w' (write): Borra el contenido existente y escribe desde cero. 'a' (append): Agrega las nuevas líneas al final del archivo. 
            writer = csv.writer(csvfile)
            writer.writerow(['-------------------','--------------','DECLARACION X BOTS (1/2)'])
            writer.writerow(['Timestamp', 'Accion', 'Detalle'])
            print(f'NDC{monitor} -> Abriendo navegador en la pagina de inicio')
            sys.stdout.flush() #permite enviar el print al frame del home
            writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), 'Navegacion', 'TG Bots -> Pagina de inicio cargada'])

        try: # Este bloque permite al RPA esperar a que el objeto o componente termine de cargar.
            if getattr(sys, 'frozen', False):
                # Si es un ejecutable, la base es la carpeta donde está el archivo .exe
                base_dir = Path(sys.executable).resolve().parent
            else:
                # Si es un script .py, la base es la carpeta donde está el archivo .py
                base_dir = Path(__file__).resolve().parent
                   
            # --- PASO NUEVO: Cargar credenciales desde el archivo ---
            keys = base_dir / 'keys.txt'
            with open(keys, 'r') as f:
                credentials = json.load(f)
          
            user_val = credentials['user']
            pass_val = credentials['pass']
            # -------------------------------------------------------

            usuario_crecencial = driver.find_element(By.XPATH, f"//*[@id='username']")
            usuario_crecencial.send_keys(user_val)
            time.sleep(2)
            
            clave_crecencial = driver.find_element(By.XPATH, f"//*[@id='password']")
            clave_crecencial.send_keys(pass_val) 
            time.sleep(2)

            btn_crecencial = driver.find_element(By.XPATH, f"//*[@id='btn_valida_ingreso']")
            btn_crecencial.click()
            driver.execute_script("document.body.style.zoom='65%'")
            print(f"NDC{monitor} -> Credenciales y Acceso ok")
            sys.stdout.flush()
            with open( archivolog, 'a', newline='') as csvfile: 
                writer = csv.writer(csvfile)
                writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), 'Credenciales', 'NDC -> Credenciales y Acceso ok'])
                writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
            time.sleep(10) # Espera para que cargue bien la pagina
    #
    #------- Importación de datos desde excel DATA.xlsx 
    #
            try:
                print(f"NDC{monitor} -> Extracción de datos desde excel")
                sys.stdout.flush()
                driver.execute_script("document.body.style.zoom='65%'")
                filesheet = "./Digitacion/DATA.xlsx"
                wb = load_workbook(filesheet, data_only=True)
                declaraciones = wb[f'DataParaBots{monitor}']
                valores_encontrados = {} # Diccionario para almacenar valores y sus celdas
            
                # 1. Recorre la columna 'A' celda por celda
                for fila in range(1, declaraciones.max_row + 1):
                    celda = declaraciones[f'A{fila}']
                    valor_celda = celda.value
            
                    # Solo procesa celdas con un valor
                    if valor_celda is not None:
                        # Si el valor ya está en el diccionario, añade la coordenada de la celda
                        if valor_celda in valores_encontrados:
                            valores_encontrados[valor_celda].append(celda.coordinate)
                        # Si es la primera vez que encuentras el valor, crea una nueva entrada
                        else:
                            valores_encontrados[valor_celda] = [celda.coordinate]
            
                # 2. Recorre el diccionario y extrae los datos de las otras columnas
                for valor_a, celdas_a in valores_encontrados.items():
                    print(f"NDC{monitor} -> registro {celdas_a} de {declaraciones.max_row}")
                    sys.stdout.flush()
                    driver.execute_script("document.body.style.zoom='65%'")
                    print(f'NDC{monitor} -> Orden Buscada: {valor_a}')
                    sys.stdout.flush() # This is crucial for real-time output
                    with open( archivolog, 'a', newline='') as csvfile: 
                        writer = csv.writer(csvfile)
                        writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' Declaracion ', f' NDC -> Orden Buscada: {valor_a}'])
            #
            # -------  OT Busqueda -------
            #
                    folio_ot = driver.find_element(By.XPATH, f"//*[@id='numero_ot']")
                    folio_ot.clear()  # limpia el contenido anterior
                    folio_ot.send_keys(valor_a)
                    folio_ot.click()
                    #print("NDC{monitor} -> registro de orden de trabajo ok")
                    time.sleep(1)

                    wait = WebDriverWait(driver, 8)  # Espera un máximo de 15 segundos la visibillidad del boton, pero si lo encuentra pasa inmediatamente
                    btn_buscar = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='boton-buscar']/span")))
                    btn_buscar.click()
                    #print("NDC{monitor} -> btn de inicio de busqueda de orden ok")
                    time.sleep(5)
            #
            # -------  OT Validación trabajo no asociado -------
            #
                    espera_alerta = WebDriverWait(driver, 2)
    
                    # Verifica si la alerta está presente
                    try:
                        espera_alerta.until(
                            EC.text_to_be_present_in_element((By.XPATH, "//*[@id='alert']"), "Existe actividad en la OT no asociada a un trabajo")
                        )
                        # Si la alerta existe, busca y haz clic en el botón
                        btn_alerta = driver.find_element(By.XPATH, "/html/body/div[8]/div[3]/div/button/span")
                        btn_alerta.click()
                        print(f"NDC{monitor} -> Alerta 'Actividad en la OT no asociada' detectada y aceptada.")
                        alerta_Ot_no_asociada = True    
                    except:
                        alerta_Ot_no_asociada = True
                        # Si la alerta no aparece en 2 segundos, no hace nada y el código continúa
                        #print(f"NDC{monitor} -> Sin alerta, continua el proceso.")

                    if alerta_Ot_no_asociada:
            #
            # -------  OT Validación trabajo ya declarado -------
            #
                        try:
                            WebDriverWait(driver, 6).until(
                                EC.any_of(  
                                   EC.text_to_be_present_in_element((By.XPATH, "//*[@id='div_declaraciones_realizadas']/table/tbody/tr[3]/td[6]"), "Declarada"),
                                   EC.text_to_be_present_in_element((By.XPATH, "//*[@id='div_declaraciones_realizadas']/table/tbody/tr[3]/td[6]"), "En proceso")
                                )
                            )
                            print(f"NDC{monitor} -> Error - Orden fue declarada con anterioridad.")
                            sys.stdout.flush()
                            print("-" * 50) # Separador para mejor lectura
                            sys.stdout.flush()
                            with open( archivolog, 'a', newline='') as csvfile: 
                                writer = csv.writer(csvfile)
                                writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' CierreOrden ', ' NDC -> NO Declarada - Orden fue declarada con anterioridad'])
                                writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
                            with open( statusorden, 'a', newline='') as csvfile: 
                                writer = csv.writer(csvfile)
                                writer.writerow([ f'{valor_a} ; Declarada o En Proceso ;', time.strftime('%Y-%m-%d %H:%M:%S')])    
                        except:
                    #
                    #--- Seleccion de la actividad realizada
                    #
                            try:
                                wait = WebDriverWait(driver, 6)
                                seleccion_trabajo = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='trabajo-material']")))
                                
                                dropdown = Select(seleccion_trabajo)  # Create a Select object for the dropdown
                                dropdown.select_by_index(1)
    
                                selected_option_text = dropdown.first_selected_option.text #Captura el dato visible del select
                                if selected_option_text == "Access Points Anexo WIFI (Extensor)" or selected_option_text == "Access Points Anexo":
                                #if selected_option_text == "Access Points Anexo WIFI (Extensor)" :    
                                    print(f"NDC{monitor} -> 'Access Point Anexo WIFI (Extensor)' se mueve a la siguente opcion.")
                                    dropdown.select_by_index(2) #Pasa a la siguiente opción
                                    time.sleep(2) 
                                else:
                                    #print(f"NDC{monitor} -> Trabajo seleccionado.")
                                    time.sleep(2) 
                                seguir_declaracion = True
                            except:
                                print(f"NDC{monitor} -> Error - Actividad no termino de cargar o existe un problema con la actividad")
                                sys.stdout.flush()
                                print("-" * 50) # Separador para mejor lectura
                                sys.stdout.flush()
                                with open( archivolog, 'a', newline='') as csvfile: 
                                    writer = csv.writer(csvfile)
                                    writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' CierreOrden ', ' NDC -> NO Declarada - Actividad no termino de cargar o existe un problema con la actividad'])
                                    writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
                                seguir_declaracion = False 
                                with open( statusorden, 'a', newline='') as csvfile: 
                                    writer = csv.writer(csvfile)
                                    writer.writerow([ f'{valor_a} ; existe un problema con la actividad ;', time.strftime('%Y-%m-%d %H:%M:%S')]) 

                    #
                    #--- Bloque de Carga de material
                    #       
                            time.sleep(8) 
                            if seguir_declaracion:
                                for coordenada in celdas_a:
                                    # Usa una expresión regular para extraer el número de la fila
                                    # Por ejemplo, de 'A5' extrae el número 5
                                    match = re.search(r'\d+', coordenada)
                                    if match:
                                        fila_numero = int(match.group())

                                        # Usa el número de fila para acceder a las celdas P y O
                                        folio = declaraciones[f'A{fila_numero}']
                                        producto = declaraciones[f'U{fila_numero}']
                                        cantidad = declaraciones[f'O{fila_numero}']

                                        # Muestra los valores de las tres celdas relacionadas
                                        #print(f"NDC{monitor} -> Material a declarar -> Fila {fila_numero}: Folio: {folio.value}, Producto: {producto.value}, Cantidad: {cantidad.value}")
                                        with open( archivolog, 'a', newline='') as csvfile: 
                                            writer = csv.writer(csvfile)
                                            writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' Declaracion ', f'NDC -> Material a declarar ->  Folio: {folio.value} , Producto: {producto.value} , Cantidad: {cantidad.value} , Fila: {fila_numero}'])
                        #
                        #------- Carga de materiales -> Detalle ------
                        #
                                        
                                        ingreso_producto = driver.find_element(By.XPATH, f"//*[@id='in_codigo']")
                                        valor_del_campo = ingreso_producto.get_attribute("value")
                                        time.sleep(2)
                                        # Verifica si el campo está vacío, esto significa que el codigo anterior fue reconocido
                                        if not valor_del_campo:
                                            ingreso_producto.send_keys(producto.value)
                                            ingreso_producto.click
                                        else:
                                            # El campor no esta vacio termina el proceso y pasa a otra orden
                                            print(f"NDC{monitor} -> Error - Codigo de material no reconocido")
                                            sys.stdout.flush()
                                            print("-" * 50) # Separador para mejor lectura
                                            sys.stdout.flush()
                                            with open( archivolog, 'a', newline='') as csvfile: 
                                                writer = csv.writer(csvfile)
                                                writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' CierreOrden ', ' NDC -> NO Declarada - Codigo de material no reconocido'])
                                                writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
                                            with open( statusorden, 'a', newline='') as csvfile: 
                                                writer = csv.writer(csvfile)
                                                writer.writerow([ f'{valor_a} ; Codigo de material no reconocido ;', time.strftime('%Y-%m-%d %H:%M:%S')]) 

                                            saltar_btn_procesar= False
                                            salto_a_otra_orden = False
                                            break                                 

                                        wait = WebDriverWait(driver, 10)  #  Espera explícita hasta que el elemento 'seleccion_trabajo' sea visible Espera un máximo de 10 segundos
                                        producto_espera_cargando_material = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='cargando_material']")))
                                        

                                        wait = WebDriverWait(driver, 10)
                                        producto_espera_activacion = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='id_material']")))
                                        
                                        time.sleep(1)
                                        ingreso_cantidad = driver.find_element(By.XPATH, f"//*[@id='cantidad']")
                                        ingreso_cantidad.send_keys(cantidad.value)
                                        ingreso_cantidad.click()
                                        #print("NDC{monitor} -> Cantidad agregada")
                                        time.sleep(1)
                                        btn_agregar_material = driver.find_element(By.XPATH, f"//*[@id='ingresar_nuevo']/span")
                                        btn_agregar_material.click()
                                        #print("NDC{monitor} -> Cantidad btn agregar material ok")
                                        time.sleep(1)
                        #
                        #------- Carga de materiales -> Alerta por material repetido ---      
                        #                           

                                        try:
                                            wait = WebDriverWait(driver, 1)  # Espera un máximo de 1x segundos la visibillidad del boton, pero si lo encuentra pasa inmediatamente
                                            btn_buscar = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[8]/div[3]/div/button/span")))
                                            btn_buscar.click()
                                            wait = WebDriverWait(driver, 1)  # Espera un máximo de x segundos la visibillidad del boton, pero si lo encuentra pasa inmediatamente
                                            btn_buscar = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='in_codigo']")))
                                            btn_buscar.clear()                                           
                                            print(f'NDC{monitor}-> Material Repetido')

                                        except:
                                            saltar_btn_procesar= True
                                            time.sleep(1)
                    
                            else:
                                salto_a_otra_orden = False
                    #
                    #--- Procesar -> PreCierr de declaración
                    #  
                            if saltar_btn_procesar: # si es True ejecuta el bloque de procesar declaración
                                btn_procesar_material = driver.find_element(By.XPATH, f"//*[@id='btn_procesar_declaracion']/span")
                                btn_procesar_material.click()
                                #print("NDC{monitor} -> btn procesar material ok")
                                time.sleep(8)
    
                                try:  # bloque que detecta error en materiales no encontrado o fuera del estandar de trabajo
                                    WebDriverWait(driver, 2).until(
                                        EC.text_to_be_present_in_element((By.TAG_NAME, 'body'), "Debe declarar material")
                                    )
                                    boton_material_notificacion = driver.find_element(By.XPATH, "/html/body/div[8]/div[3]/div/button/span") # Usar XPATH Procesar
                                    boton_material_notificacion.click()
                                    print(f"NDC{monitor} -> Error - no detecta producto erroneo.")
                                    sys.stdout.flush()
                                    print("-" * 50) # Separador para mejor lectura
                                    sys.stdout.flush()
                                    with open( archivolog, 'a', newline='') as csvfile: 
                                        writer = csv.writer(csvfile)
                                        writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' CierreOrden ', ' NDC -> NO Declarada - producto erroneo'])
                                        writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
                                    with open( statusorden, 'a', newline='') as csvfile: 
                                        writer = csv.writer(csvfile)
                                        writer.writerow([ f'{valor_a} ; producto erroneo ;', time.strftime('%Y-%m-%d %H:%M:%S')])     
                                    salto_a_otra_orden = False
                                except:
                                    salto_a_otra_orden = True
                    #
                    #--- Procesar -> Validación de stock final
                    #
                            if salto_a_otra_orden: # si es True ejecuta el bloque de validacion antes de procesar ------
                                time.sleep(3)
                                try:
                                    WebDriverWait(driver, 6).until(
                                        #EC.text_to_be_present_in_element((By.TAG_NAME, 'body'), "Error Stock") # no tocar
                                        EC.text_to_be_present_in_element((By.TAG_NAME, 'body'), "sin stock suficiente")
                                    )
                                    print(f"NDC{monitor} -> No Declarada - Stock InSuficiente.")
                                    sys.stdout.flush()
                                    print("-" * 50) # Separador para mejor lectura
                                    sys.stdout.flush()
                                    with open( archivolog, 'a', newline='') as csvfile: 
                                        writer = csv.writer(csvfile)
                                        writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' CierreOrden ', ' NDC -> NO Declarada - Stock InSuficiente'])
                                        writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
                                    with open( statusorden, 'a', newline='') as csvfile: 
                                                writer = csv.writer(csvfile)
                                                writer.writerow([ f'{valor_a}; falta-stock ;', time.strftime('%Y-%m-%d %H:%M:%S')]) 
                                    time.sleep(3)
                                    try:
                                       boton_aceptar = WebDriverWait(driver, 6).until(
                                           EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Aceptar')]"))
                                       )
                                       
                                       # 2. Una vez encontrado, hacer el clic
                                       boton_aceptar.click()
                                       print("Se encontro el boton y se hizo clic en Aceptar.")                                      
                                    except:
                                       ActionChains(driver).send_keys(Keys.TAB).perform()
                                       ActionChains(driver).send_keys(Keys.ENTER).perform()
                                       print("Click en 'Tab-action' realizado con éxito.")
                                       sys.stdout.flush()
                                    
                                    time.sleep(3)
                                    boton_cerrar = driver.find_element(By.XPATH, "/html/body/div[6]/div[3]/div/button[2]") 
                                    boton_cerrar.click()
                                except:
                                        print(f"NDC{monitor} -> Stock correcto se procede a procesar.")
                                        try:
                                            boton_procesar2 = WebDriverWait(driver, 8).until(
                                                EC.element_to_be_clickable((By.XPATH, "/html/body/div[7]/div[3]/div/button[1]"))
                                            )

                                            # 2. Ahora que lo tenemos guardado en 'boton_procesar', hacemos el click
                                            boton_procesar2.click()
                                            print(f"NDC{monitor} -> apunto de procesar - boton procesar clickable.")
                                        except:
                                            WebDriverWait(driver, 8).until(
                                            EC.text_to_be_present_in_element((By.TAG_NAME, 'body'), "Procesar")
                                            )
                                            print(f"NDC{monitor} -> apunto de procesar Declarada - boton procesar in_element.")
                                            time.sleep(10)
                                        with open( archivolog, 'a', newline='') as csvfile: 
                                            writer = csv.writer(csvfile)
                                            writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' CierreOrden ', ' NDC -> Declarada sin refuerzo'])
                                            print(f"NDC{monitor} -> apunto de procesar Declarada - boton xpath")
                                        try:
                                            time.sleep(10)
                                            btn_procesar_notificacion = driver.find_element(By.XPATH, f"/html/body/div[9]/div[3]/div/button/span")
                                            btn_procesar_notificacion.click()
                                            time.sleep(10)
                                            print(f"NDC{monitor} -> Material declarado correctamente.")
                                            sys.stdout.flush()
                                            print("-" * 50) # Separador para mejor lectura
                                            sys.stdout.flush()
                                            with open( archivolog, 'a', newline='') as csvfile: 
                                                writer = csv.writer(csvfile)
                                                writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' CierreOrden ', ' NDC -> SI Declarada >>>>>'])
                                                writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
                                            with open( statusorden, 'a', newline='') as csvfile: 
                                                writer = csv.writer(csvfile)
                                                writer.writerow([ f'{valor_a}; Declarada o En Proceso ;', time.strftime('%Y-%m-%d %H:%M:%S')]) 
                                        except: # Se mueve a otra pestaña si el procesamiento final se demora mas de lo esperado.
                                            print(f"NDC{monitor} -> Orden no declarada, fallo la ultima notificación.")
                                            sys.stdout.flush()
                                            print("-" * 50) # Separador para mejor lectura
                                            sys.stdout.flush()
                                            with open( archivolog, 'a', newline='') as csvfile: 
                                                writer = csv.writer(csvfile)
                                                writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' CierreOrden ', ' NDC -> Orden no declarada, fallo la ultima notificación.'])
                                                writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
                                             # Obtener la ventana actual (la pestaña principal)
                                            original_window = driver.current_window_handle
                                            # Abrir una nueva pestaña usando JavaScript
                                            # Esto creará una nueva pestaña, pero el driver seguirá enfocado en la original
                                            driver.execute_script("window.open('');")
                                            # Esperar a que la nueva ventana se abra (es una buena práctica)
                                            time.sleep(4)
                                            # Obtener todas las ventanas/pestañas abiertas
                                            #all_windows = driver.window_handles
                                            # Buscar la nueva ventana y cambiar el enfoque a ella

                                            # 4. Cambiar el enfoque a la ÚLTIMA ventana abierta
                                            # [-1] selecciona el último elemento de la lista de handles
                                            driver.switch_to.window(driver.window_handles[-1])
                                           # for window_handle in all_windows:
                                           #     if window_handle != original_window:
                                           #         driver.switch_to.window(window_handle)
                                           #         break
                                            driver.get("http://eps.vtr.cl/eps/declaracion_consumo.do#")
                                            # Si quieres volver a la pestaña original más tarde, simplemente haz esto:
                                            # driver.switch_to.window(original_window)

                    else:
                        print(f"NDC{monitor} -> Error - OT no asociada a un trabajo.")
                        sys.stdout.flush()
                        print("-" * 50) # Separador para mejor lectura
                        sys.stdout.flush()
                        with open( archivolog, 'a', newline='') as csvfile: 
                            writer = csv.writer(csvfile)
                            writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' CierreOrden ', ' NDC -> NO Declarada - Existe actividad en la OT no asociada a un trabajo'])
                            writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
            #
            #--- Eliminación -> Campos correctamente procesados, serán eliminados
            #               
                for valor_a, celdas_a in valores_encontrados.items():
                      print(f"filas a eliminar{celdas_a}, esta seran eliminadas")
        #
        #--- Errores -> de extraccion de datos
        #
            except FileNotFoundError:
                print(f"Error: El archivo '{filesheet}' no se encuentra en el directorio especificado.")
            except KeyError:
                print("Error: La hoja de calculo 'Hoja 2' no existe en el archivo.")
            except Exception as e:
                with open( archivolog, 'a', newline='') as csvfile: 
                                           writer = csv.writer(csvfile)
                                           writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' CierreOrden ', ' TG Bots -> Error: Ocurrió un error inesperado.'])
                                           writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
                print(f"Error: Ocurrió un error inesperado: {e}")
    #        
    #------- TG Bots Final del proceso ya sea por 100% de registros procesados o error sin alternativa de continuar ------
    #
        except Exception as e:
            time.sleep(5)
            with open( archivolog, 'a', newline='') as csvfile: 
                                           writer = csv.writer(csvfile)
                                           writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' CierreOrden ', ' TG Bots -> No se pudo encontrar o interactuar con el elemento.'])
                                           writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
            self.fail(f"No se pudo encontrar o interactuar con el elemento: {e}")
        finally:
            if 'wb' in locals() and wb:
                wb.close()
                print(f"NDC{monitor} -> El archivo de Excel ha sido cerrado.")
                with open( archivolog, 'a', newline='') as csvfile: 
                                           writer = csv.writer(csvfile)
                                           writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' Navegación ', ' TG Bots -> El archivo de Excel ha sido cerrado.'])
                                           writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
       

    
    def tearDown(self):
        print(f"NDC{monitor} -> Cierre")
        sys.stdout.flush()
        with open( archivolog, 'a', newline='') as csvfile: 
                                           writer = csv.writer(csvfile)
                                           writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), ' Navegación ', ' TG Bots -> Cierre'])
                                           writer.writerow(['-------------------','--------------','--------------------------------------------------------------------------------'])
        time.sleep(2)
        self.driver.close()

if __name__ == '__main__':
    unittest.main()  



'''
      #------- Ejercicio 21------

      #      wb = None 
      #      filesheet = "./Digitacion/DATA.xlsx"
      #      
      #      try:
      #          wb = load_workbook(filesheet)
      #          hojas = wb.sheetnames
      #          print(hojas)
      #      
      #          declaraciones = wb['Hoja 2']
      #      
      #          for i in range(1, 15):
      #              # Lee los datos de las columnas no correlativas
      #              folio = declaraciones[f'A{i}']
      #              producto = declaraciones[f'P{i}']
      #              cantidad = declaraciones[f'O{i}']
      #      
      #              # Imprime los valores
      #              print(f"Folio: {folio.value}, Producto: {producto.value}, Cantidad: {cantidad.value}")
      #              
      #      except FileNotFoundError:
      #          print(f"Error: El archivo '{filesheet}' no se encuentra en el directorio especificado.")
      #      except KeyError:
      #          print("Error: La hoja de cálculo 'Hoja 2' no existe en el archivo.")
      #      except Exception as e:
      #          print(f"Ocurrió un error inesperado: {e}")
      #      finally:
      #          # Cierra el archivo solo si fue abierto exitosamente
      #          if wb is not None:
      #              wb.close()
      #              print("El archivo de Excel ha sido cerrado.")
      #      writer.writerow([time.strftime('%Y-%m-%d %H:%M:%S'), 'Extraccion datos exel', 'Datos Cargados a las variables'])



#Ejercicio 20 bloqueo de notificaciones emergentes de una pagina web

opciones = Options()
 # Enviar Argumento (1 permitiendo la notificación, 2 bloquea la notificación)

opciones.add_experimental_option("prefs",{
  "profile.default_content_setting_values.notifications" : 2
})

driver = webdriver.Chrome(options=opciones)
driver.get("https://developer.mozilla.org/es/docs/Web/API/notificacion")
time.sleep(8)
driver.close()



#Ejercicio 19 Click sobre un mensaje emergente de javascript

driver = webdriver.Chrome()
driver.get("file:///C:/Users/Jose%20Huaiquiman/Desktop/MisScript/Priv_Selenium_Python/alert_simple.html")
time.sleep(3)

# --------- Alerta Simple  ----------
alerta_simple = driver.find_element(By.NAME, "alert")
alerta_simple.click()
time.sleep(3)

try:
    # La forma correcta y moderna de cambiar a una alerta
    alerta_js = driver.switch_to.alert

    # Opcionalmente, puedes obtener el texto de la alerta
    mensaje_alerta = alerta_js.text
    print(f"Mensaje de la alerta: {mensaje_alerta}")

    # Aceptar la alerta (hacer clic en 'Aceptar' o 'OK')
    alerta_js.accept()


except Exception as e:
    print(f"No se pudo cambiar a la alerta. Posiblemente no está presente. {e}")
time.sleep(3)

# --------- Alerta Prompt  ----------
alerta_simple_prompt = driver.find_element(By.NAME, "prompt_alert")
alerta_simple_prompt.click()
time.sleep(3)

try:
    # La forma correcta y moderna de cambiar a una alerta
    alerta_js = driver.switch_to.alert

    # Opcionalmente, puedes obtener el texto de la alerta
    mensaje_alerta = alerta_js.text
    print(f"Mensaje de la alerta: {mensaje_alerta}")

    # Aceptar la alerta (hacer clic en 'Aceptar' o 'OK')
    #alerta_js.accept()  # Boton Aceptar
    alerta_js.dismiss()  # Boton Cancelar

except Exception as e:
    print(f"No se pudo cambiar a la alerta. Posiblemente no está presente. {e}")
time.sleep(3)
driver.close()




#Ejercicio 18 extraer datos de una tabla y almacenar en una variable.

driver = webdriver.Chrome()
driver.get("https://www.w3schools.com/html/html_tables.asp")
time.sleep(5)

valor = driver.find_element(By.XPATH,"//*[@id='customers']/tbody/tr[2]/td[2]").text
print(f"El valor encontrado es: {valor}")

rows=driver.find_elements(By.XPATH,"//*[@id='customers']/tbody/tr") #rescata la cantidad de filas
num_rows = len(rows)
col=driver.find_elements(By.XPATH,"//*[@id='customers']/tbody/tr[1]/th") #rescata la cantidad de columnas
num_cols =len(col)
print(num_rows)
print(num_cols)

for n in range(2,  num_rows+1):
    for b in range(1, num_cols+1):
        dato = driver.find_element(By.XPATH,"//*[@id='customers']/tbody/tr["+str(n)+"]/td["+str(b)+"]").text
        print(dato, end='  ;  ')
    print()

    

#Ejercicio 17  Corre una aplicacion Python en segundo plano

# Instrucción para correr Chrome en segundo plano esperando los resultados
# Opciones para correr Chrome en modo headless
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-dev-shm-usage") # Opcional, pero recomendado en algunos entornos

# Inicializa el driver con las opciones para que corra en segundo plano
driver = webdriver.Chrome(options=chrome_options)
driver.get("https://www.google.com")

# Esperamos a que la página cargue completamente
time.sleep(5)

# Encontramos la barra de búsqueda por su nombre (generalmente 'q')
busqueda = driver.find_element(By.NAME, "q")
busqueda.send_keys("huaiqui")
#busqueda.send_keys(webdriver.common.keys.Keys.RETURN) # Presiona Enter para iniciar la búsqueda

# Esperamos a que los resultados de la búsqueda aparezcan
time.sleep(5)

# Bucle para encontrar y mostrar los 10 primeros resultados de la búsqueda
# Utiliza un XPath más genérico para mayor estabilidad
for i in range(1, 6):
    try:
        # Este XPath busca los títulos de los resultados de búsqueda.
        # Puede que necesites ajustarlo si el HTML de Google cambia.
       #link_element = driver.find_element(By.XPATH, f"//*[@id='input']").text
        link_element = driver.find_element(By.XPATH, f"(//h3)[{i}]")
        #print(f"Resultado {i}: {link_element.text}")
        
          # Obtenemos el texto del elemento
        link_text = link_element.text
        
        # Imprimimos el texto
        print(f"Resultado {i}: {link_text}")
    except Exception as e:
        print(f"No se encontró el resultado {i}. Puede que no existan más resultados. {e}")
        break  # Salimos del bucle si no se encuentra un resultado

driver.close()
#/html/body/ntp-app//div/div[2]/cr-searchbox//div/cr-searchbox-dropdown//div/div[1]/div/cr-searchbox-match[4]//div/div[2]/span[3]/span[2]
#//*[@id="input"]




# Crear Ejecutable de python
https://www.youtube.com/watch?v=vaKjEX_8Eqk&list=PLas30d-GGNa2UW9-1H-NCNrUocvWD9cyh&index=22

Excelente video para dejar tareas programables.




#Ejercicio 16

driver = webdriver.Chrome()
driver.get("https://www.w3schools.com/html/default.asp")
time.sleep(5)

# content=driver.find_element_by_css_selector('a.w3-bue') instruccion esta obsoleta
#content = driver.find_element(By.CSS_SELECTOR, 'a.w3-bue')
#content.click()

try:
    content = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "a.w3-bue"))
    )
    print("¡Elemento encontrado!")
except:
    print("Error: No se encontró el elemento dentro del tiempo límite.")

# Si se encontró el elemento, puedes continuar con la interacción
content.click() # Por ejemplo

driver.quit()




#Ejercicio 15 Encontar un link tipo hipertexto

driver = webdriver.Chrome()
driver.get("https://www.w3schools.com")
time.sleep(3)
encontrar_link = driver.find_element(By.LINK_TEXT, "Learn PHP")
encontrar_link.click()





#Ejercicio 14  validación de carga de elemento o elemento disponible

driver = webdriver.Chrome()
driver.get("https://google.com")
time.sleep(3)
displalemen = driver.find_element(By.NAME, "btnI")
print(displalemen.is_displayed())# regresa true o false si ya cargo el elemento
print(displalemen.is_enabled())# regresa un true o false si el elemento esta disponible





# Ejercicio 13  Metodo que simula el movimiento del mouse sobre un link
driver = webdriver.Chrome()
driver.get("https://google.com")
time.sleep(3)
#elements = driver.find_element_by_link_text("Privacidad")  Metodo obsoleto
elements = driver.find_element(By.LINK_TEXT, "Privacidad")

hover = ActionChains(driver).move_to_element(elements)
hover.perform()
time.sleep(5) 

driver.quit()




# Ejercicio 12 seleccionar y cambiar de estado los radio boton

class usando_unittest(unittest.TestCase):
    
    def setUp(self):
        self.driver = webdriver.Chrome()

    def test_usando_radio_bt(self):
        driver = self.driver

        driver.get("https://www.w3schools.com/howto/howto_css_custom_checkbox.asp")
        time.sleep(2)

        rabio_boton_a = driver.find_element(By.XPATH,"//*[@id='main']/div[3]/div[1]/input[3]")
        rabio_boton_a.click()
        time.sleep(2)

        rabio_boton_b = driver.find_element(By.XPATH,"//*[@id='main']/div[3]/div[1]/input[4]")
        rabio_boton_b.click()
        time.sleep(2)
       
        rabio_boton_c = driver.find_element(By.XPATH,"//*[@id='main']/div[3]/div[1]/input[3]")
        rabio_boton_c.click()
        time.sleep(2)
       
    def tearDown(self):
        self.driver.close()

if __name__ == '__main__':
    unittest.main() 

    


# Ejercicio 11 seleccionar una de las opciones del select o recorrerlo

class usando_unittest(unittest.TestCase):
    
    def setUp(self):
        self.driver = webdriver.Chrome()

    def test_usando_select(self):
        driver = self.driver
        driver.get("https://www.w3schools.com/howto/howto_custom_select.asp")
        time.sleep(5)
        select_element = driver.find_element(By.XPATH, "//*[@id='main']/div[3]/div[1]/select")
        opcion_list = select_element.find_elements(By.TAG_NAME,"option")
        time.sleep(5)
        # En este bloque recorre el select
        for option_individual in opcion_list:
            print("Los valores son: %s" % option_individual.get_attribute("value"))
            option_individual.click()
            time.sleep(3)
        # En este bloque lo va a buscar de manera directa   
        seleccionar = Select(driver.find_element(By.XPATH, "//*[@id='main']/div[3]/div[1]/select"))
        seleccionar.select_by_value("10")
        time.sleep(3)
        driver.quit()
       
    def tearDown(self):
        self.driver.close()

if __name__ == '__main__':
    unittest.main()  


    



#Ejercicio 10 Uso de click en toggle o swith

class usando_unittest(unittest.TestCase):
    
    def setUp(self):
        self.driver = webdriver.Chrome()

    def test_usando_togle(self):
        driver = self.driver

        driver.get("https://www.w3schools.com/howto/howto_css_switch.asp")
        time.sleep(5)
        toggle = driver.find_element(By.XPATH, "//*[@id='main']/label[3]/div")
        toggle.click()
        time.sleep(5)
        toggle.click()
        time.sleep(5)
    
       
    def tearDown(self):
        self.driver.close()

if __name__ == '__main__':
    unittest.main()  

    


#Ejercicio 9 comparando imagenes con OpenCV, screenshot.

class usando_unittest(unittest.TestCase): #Importante el TestCase ejecuta las funciones en orden alfabetico
    
    def setUp(self):
        self.driver = webdriver.Chrome()

    def test_a_usando_opencv(self):  
        driver = self.driver
        driver.get("http://www.google.com")
        
        #permite sacar una foto de pantalla de la pagina que esta abierta
        # la cual compara con una imagen guarda y para esto obviamente se necesita una imagen previamente guardada
        # se recomiendo .png por su mejor resolución
        driver.save_screenshot('img2.png') 
        time.sleep(3)

    def test_b_comparando_imagenes(self):
        img1=cv2.imread('img1.png') # se carga la imagen a una variable
        img2=cv2.imread('img2.png') # se carga la imagen a una variable
        if img1 is None or img2 is None:
            print("Error: No se pudieron cargar una o ambas imágenes. Verifica que la ruta y el nombre del archivo sean correctos.")
            return # Termina la función si hay un problema con la carga de las imágenes
        try:
            diferencia = cv2.absdiff(img1, img2) #aqui se realiza la comparación de imagenes y guarda el resultado en variable diferencia
            imagen_gris = cv2.cvtColor(diferencia, cv2.COLOR_BGR2GRAY) # esta instrucción con atributo COLOR_BGR2GRAY es un indicador para que realice todo en escala de grises.
            contours,_ = cv2.findContours(imagen_gris, cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE) # " contours,_ ""Asigna el primer valor que la función devuelva a la variable contours, y descarta el segundo valor, ya que no lo voy a utilizar en mi código."
            #El Significado del Guion Bajo (_) En Python, un guion bajo simple (_) se usa como un nombre de variable para un valor que no te interesa o que estás decidiendo ignorar intencionalmente. Es una señal para otros programadores de que no planeas usar esa parte del resultado de la función.
            
            # Puedes agregar más código aquí para trabajar con los contornos
            print("Las imagenes se compararon con exito..!")
        except cv2.error as e:
            print(f"Ocurrió un error de OpenCV: {e}")
        except Exception as e:
            print(f"Ocurrió un error inesperado: {e}")
    
        for c in contours:
            if cv2.contourArea(c) >=20: # aqui realiza una comparación y si el rango es mayor a 20 existen diferencias
                posicion_x, posicion_y, ancho, alto = cv2.boundingRect(c) #identifica la diferencia y la guarda las dimensiones en las variables.
                cv2.rectangle(img1,(posicion_x, posicion_y),(posicion_x+ancho, posicion_y+alto),(0,0,255),2)

        while(1):
            cv2.imshow('Imagen1', img1) #Esta instrucción muestra las imagenes.
            cv2.imshow('Imagen2', img2)
            cv2.imshow('diferencia detectadas', diferencia)
            teclado = cv2.waitKey(5) & 0xFF #rutina exclusiva de opencv
            if teclado == 27:
                break
        cv2.destroyAllWindows()
    def tearDown(self):
        self.driver.close()

if __name__ == '__main__':
    unittest.main()  
 

    

#Ejercicio 8 esperar que un objeto o componente termine de cargar.

class usando_unittest(unittest.TestCase):
    
    def setUp(self):
        self.driver = webdriver.Chrome()

    def test_usando_implicit_wait(self):
        driver = self.driver

        driver.implicitly_wait(5) #Espera 5 segundos mientras buscas el objeto y si no lo encuentras se sale, especial para encontrar componentes en carrucel de imagenes por ejemplo
        driver.get("https://www.google.com")
        myDynamicElement = driver.find_element(By.NAME, "q")

    def tearDown(self):
        self.driver.close()

if __name__ == '__main__':
    unittest.main()  

    

#Ejercicio 7 "explicit_wait" esperar que un objeto o componente termine de cargar.

class usando_unittest(unittest.TestCase):
    
    def setUp(self):
        self.driver = webdriver.Chrome()

    def test_usando_explicit_wait(self):
        driver = self.driver
        driver.get("https://www.google.com")

        try:             # Este bloque permite esperar a que el objeto o componente termine de cargar.
            element = WebDriverWait(driver,10).until(EC.presence_of_element_located((By.NAME,"q")))
        finally:
            driver.quit()

    def tearDown(self):
        self.driver.close()

if __name__ == '__main__':
    unittest.main()  


    

#Ejercicio 6 volver atras.

class usando_unittest(unittest.TestCase):
    
    def setUp(self):
        self.driver = webdriver.Chrome()

    def test_pagina_siguiente_o_anterior(self):
        driver = self.driver

        driver.get("https://www.gmail.com")
        time.sleep(5)

        driver.get("https://www.google.com")
        time.sleep(5)
'''
